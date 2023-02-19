from tkinter import *
from tkcalendar import *
from tkinter import ttk, messagebox
import tkinter as tk
import openpyxl,xlrd
from openpyxl.workbook import Workbook

import pathlib

class FormulaireExcel :
    def __init__(self, root):
        self.root=root
        self.root.title("Formulaire reparation")
        self.root.geometry("575x575+300+100")

        frame1 = Frame(self.root, bg="blue")
        frame1.place(x=40, y=50, height=500,width=500)


        title= Label(frame1,text=" Departement reparation ",font=( "Algerian",20,"bold") , bg= "blue" , fg="white").place ( x= 50 , y= 30 )

        text_Produitm= Label(frame1,text=" Produit ",font=( "Algerian",10,"bold") , bg= "blue" , fg="white").place ( x= 50 , y= 100 )
        self.ecri_Produit= Entry( frame1, font=("time new roman", 15), bg= "White")
        self.ecri_Produit.place( x= 180 , y= 100)
        text_SN= Label(frame1,text=" SN ",font=( "Algerian",10,"bold") , bg= "blue" , fg="white").place ( x= 50 , y= 140 )
        self.ecri_SN= Entry( frame1, font=("time new roman", 15), bg= "White")
        self.ecri_SN.place( x= 180 , y= 140)
        text_TEST= Label(frame1,text=" TEST ",font=( "Algerian",10,"bold") , bg= "blue" , fg="white").place ( x= 50 , y= 180 )
        self.ecri_TEST= Entry( frame1, font=("time new roman", 15), bg= "White")
        self.ecri_TEST.place( x= 180 , y= 180)
        text_FAIL= Label(frame1,text="FAIL",font=( "Algerian",10,"bold") , bg= "blue" , fg="white").place ( x= 50 , y= 220 )
        self.ecri_FAIL= Entry( frame1, font=("time new roman", 15), bg= "White")
        self.ecri_FAIL.place( x= 180 , y= 220)
        text_Cause= Label(frame1,text=" Cause ",font=( "Algerian",10,"bold") , bg= "blue" , fg="white").place ( x= 50 , y=260)
        self.ecri_Cause= Entry( frame1, font=("time new roman", 15), bg= "White")
        self.ecri_Cause.place( x= 180 , y= 260)
        text_action= Label(frame1,text=" Action ",font=( "Algerian",10,"bold") , bg= "blue" , fg="white").place ( x= 50 , y= 300 )
        self.ecri_Action= ttk.Combobox(  frame1,state=" readonly",  font=("time new roman", 15))
        self.ecri_Action["values"] = (" CARTE REPARER", "CARTE NON CONFORM")
        self.ecri_Action.set("Selection")
        self.ecri_Action.place( x = 180 , y = 300)
        text_date= Label(frame1,text=" Date  ",font=( "Algerian",10,"bold") , bg= "blue" , fg="white").place ( x= 50 , y= 340 )
        self.ecri_Date= DateEntry( frame1, font=("time new roman", 15), bg= "White", date_pattern = "dd/mm/yy")
        self.ecri_Date.place( x= 180 , y= " 340")

        b1 = Button(frame1,text="Valider",font=( "new roman",15) , bg= "limegreen",bd= 5,command=self.valide).place(x = 80 , y= 400, width= 150)
        b2 = Button(frame1,text="Reinitialiser",font=( "new roman",15) , bg= "limegreen",bd= 5,command=self.reset).place(x =   300 , y= 400, width= 150)

        ####3 CRER UN FICHIER EXCEL
        fichier = pathlib.Path( r"C:\Users\HOUCINE\tunisians-loves-programming\FORMILAIRE EXCEL\BASE_REPAIR.xlsx" )
        if fichier.exists():
            pass
        else:
            fichier = Workbook ()
            sheet = fichier.active
            sheet ["A1"] = "Produit"
            sheet["B1"] =  " SN"
            sheet["C1"] =  " Test"
            sheet["D1"] = "FAIL "
            sheet["E1"] = "CAUSE "
            sheet["F1"] = " Date "
            sheet["G1"] = " Action "

            fichier.save( r"C:\Users\HOUCINE\tunisians-loves-programming\FORMILAIRE EXCEL\BASE_REPAIR.xlsx" )

    def valide(self):
        Produit = self.ecri_Produit.get()
        SN = self.ecri_SN.get()
        Test = self.ecri_TEST.get()
        FAIL = self.ecri_FAIL.get()
        CAUSE = self.ecri_Cause.get()
        Date = self.ecri_Date.get()
        Action = self.ecri_Action.get()

        fichier = openpyxl.load_workbook(
            r"C:\Users\HOUCINE\tunisians-loves-programming\FORMILAIRE EXCEL\BASE_REPAIR.xlsx")
        sheet = fichier.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=Produit)
        sheet.cell(column=2, row=sheet.max_row, value=SN)
        sheet.cell(column=3, row=sheet.max_row, value=Test)
        sheet.cell(column=4, row=sheet.max_row, value=FAIL)
        sheet.cell(column=5, row=sheet.max_row, value=CAUSE)
        sheet.cell(column=6, row=sheet.max_row, value=Date)
        sheet.cell(column=7, row=sheet.max_row, value=Action)

        fichier.save(r"C:\Users\HOUCINE\tunisians-loves-programming\FORMILAIRE EXCEL\BASE_REPAIR.xlsx")
        messagebox.showinfo("good", "STOKED DATA")
        self.reset()

    def reset(self) :
           self.ecri_Produit.delete(0,END)
           self.ecri_TEST.delete(0,END)
           self.ecri_SN.delete(0,END)
           self.ecri_Cause.delete(0,END)
           self.ecri_FAIL.delete(0,END)



                





root=tk.Tk()
obj=FormulaireExcel(root)
root.mainloop()
